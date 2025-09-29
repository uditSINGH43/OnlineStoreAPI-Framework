import json
import csv
from openpyxl import load_workbook


def read_excel_data(filepath, sheetname):
    """Excel Data Provider"""
    wb = load_workbook(filepath)  # Load the Excel workbook.
    sheet = wb[sheetname]  # Access the specified sheet by name.
    data = []  # Empty list to store data rows
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Loop through all rows from the 2nd
        # row onwards (to skip the header)
        data.append(row)  # Add each row (as a tuple) to the list
    return data  # Return list of tuples


def read_json_data(filepath):
    """JSON Data Provider"""
    with open(filepath, 'r') as f:
        data_list = json.load(f)  # List of dicts
    return [(item,) for item in data_list]  # Wrap each dict as a tuple


def read_csv_data(filepath):
    """CSV Data Provider"""
    data = []  # List to hold CSV data
    with open(filepath, 'r') as f:
        reader = csv.reader(f)  # Create a CSV reader object to read each row line by line in CSV file
        next(reader)  # Skip the header row
        for row in reader:
            data.append(tuple(row))  # Convert each row to a tuple and add to list
    return data  # Return list of tuples
